import unicodedata
import warnings
import logging
import re
import argparse

import sbol3
import openpyxl
import tyto
from .helper_functions import toplevel_named

BASIC_PARTS_COLLECTION = 'BasicParts'
COMPOSITE_PARTS_COLLECTION = 'CompositeParts'
LINEAR_PRODUCTS_COLLECTION = 'LinearDNAProducts'
FINAL_PRODUCTS_COLLECTION = 'FinalProducts'

def expand_configuration(values: dict) -> dict:
    """
    Initialize sheet configuration dictionary
    :param values: Dictionary of overrides for defaults
    :return configuration with all defaults filled in
    """
    # set up the default values
    default_values = {
        'basic_sheet': 'Basic Parts',
        'basic_parts_name': 'B1',
        'basic_parts_description': 'A11',
        'basic_first_row': 20,
        'basic_name_col': 0,
        'basic_role_col': 1,
        'basic_notes_col': 2,
        'basic_description_col': 4,
        'basic_source_prefix_col': 5,
        'basic_source_id_col': 6,
        'basic_final_col': 9,
        'basic_circular_col': 10,
        'basic_length_col': 11,
        'basic_sequence_col': 12,

        'composite_sheet': 'Composite Parts',
        'composite_parts_name': 'B1',
        'composite_parts_description': 'A11',
        'composite_first_row': 24,
        'composite_name_col': 0,
        'composite_notes_col': 1,
        'composite_description_col': 2,
        'composite_final_col': 3,
        'composite_strain_col': 4,
        'composite_context_col': 5,
        'composite_constraints_col': 6,
        'composite_first_part_col': 7,

        'sources_sheet': 'data_source',
        'sources_first_row': 2,
        'source_name_col': 1,
        'source_uri_col': 2,
        'source_literal_col': 6
    }
    # override with supplied values
    values_to_use = default_values
    if values is not None:
        for k, v in values.items():
            if k not in default_values:
                raise ValueError(f'Sheet configuration has no setting "{k}"')
            values_to_use[k] = v
    # initialize the dictionary
    return values_to_use


# TODO: remove after resolution of https://github.com/SynBioDex/pySBOL3/issues/191
def string_to_display_id(name):
    def sanitize_character(c):
        replacements = {' ': '_', '-': '_', '.': '_'}
        c = replacements.get(c, c)  # first, see if there is a wired replacement
        if c.isalnum() or c == '_':  # keep allowed characters
            return c
        else:  # all others are changed into a reduced & compatible form of their unicode name
            return f'_{unicodedata.name(c).replace(" SIGN","").replace(" ","_")}'

    # make replacements in order to get a compliant displayID
    display_id = "".join([sanitize_character(c) for c in name.strip()])
    # prepend underscore if there is an initial digit
    if display_id[0].isdigit():
        display_id = "_"+display_id
    return display_id


def read_metadata(wb: openpyxl.Workbook, doc: sbol3.Document, config: dict):
    """
    Extract metadata and build collections
    :param wb: Excel workbook to extract material from
    :param doc: SBOL document to build collections in
    :param config: dictionary of sheet parsing configuration variables
    :return: Tuple of SBOL collections for basic, composite, linear, and final parts
    """
    # Read the metadata
    ws_b = wb[config['basic_sheet']]
    bp_name = ws_b[config['basic_parts_name']].value
    bp_description = ws_b[config['basic_parts_description']].value

    ws_c = wb[config['composite_sheet']]
    if config['composite_parts_name']:
        cp_name = ws_c[config['composite_parts_name']].value
        cp_description = ws_c[config['composite_parts_description']].value
    else:
        cp_name = bp_name
        cp_description = bp_description

    # Make the collections
    basic_parts = sbol3.Collection(BASIC_PARTS_COLLECTION, name=bp_name, description=bp_description)
    doc.add(basic_parts)

    composite_parts = sbol3.Collection(COMPOSITE_PARTS_COLLECTION, name=cp_name, description=cp_description)
    doc.add(composite_parts)

    linear_products = sbol3.Collection(LINEAR_PRODUCTS_COLLECTION, name='Linear DNA Products',
                                       description='Linear DNA constructs to be fabricated')
    doc.add(linear_products)

    final_products = sbol3.Collection(FINAL_PRODUCTS_COLLECTION, name='Final Products',
                                      description='Final products desired for actual fabrication')
    doc.add(final_products)

    # also collect any necessary data tables from extra sheets
    source_table = {row[config['source_name_col']].value:row[config['source_uri_col']].value
                    for row in wb[config['sources_sheet']].iter_rows(min_row=config['sources_first_row'])
                    if row[config['source_literal_col']].value}

    # return the set of created collections
    return basic_parts, composite_parts, linear_products, final_products, source_table

# TODO: remove kludge after resolution of https://github.com/SynBioDex/tyto/issues/21
tyto_cache = {}
def tyto_lookup_with_caching(term: str) -> str:
    if term not in tyto_cache:
        try:
            tyto_cache[term] = tyto.SO.get_uri_by_term(term)
        except LookupError as e:
            tyto_cache[term] = e
    if isinstance(tyto_cache[term], LookupError):
        raise tyto_cache[term]
    else:
        return tyto_cache[term]


def row_to_basic_part(doc: sbol3.Document, row, basic_parts: sbol3.Collection, linear_products: sbol3.Collection,
                      final_products: sbol3.Collection, config: dict, source_table: dict):
    """
    Read a row for a basic part and turn it into SBOL Component
    :param doc: Document to add parts to
    :param row: Excel row to be processed
    :param basic_parts: collection of parts to add to
    :param linear_products: collection of linear parts to add to
    :param final_products: collection of final parts to add to
    :param config: dictionary of sheet parsing configuration variables
    :param source_table: dictionary mapping source names to namespaces
    :return: None
    """
    # Parse material from sheet row
    name = row[config['basic_name_col']].value
    if name is None:
        return  # skip lines without names
    try:
        raw_role = row[config['basic_role_col']].value  # look up with tyto; if fail, leave blank or add to description
        role = (tyto_lookup_with_caching(raw_role) if raw_role else None)
    except LookupError:
        logging.warning(f'Role "{raw_role}" could not be found in Sequence Ontology')
        role = None
    design_notes = (row[config['basic_notes_col']].value if row[config['basic_notes_col']].value else "")
    description = (row[config['basic_description_col']].value if row[config['basic_description_col']].value else "")
    source_prefix = row[config['basic_source_prefix_col']].value
    source_id = row[config['basic_source_id_col']].value
    final_product = row[config['basic_final_col']].value  # boolean
    circular = row[config['basic_circular_col']].value  # boolean
    length = row[config['basic_length_col']].value
    raw_sequence = row[config['basic_sequence_col']].value
    sequence = (None if raw_sequence is None else "".join(unicodedata.normalize("NFKD", raw_sequence).upper().split()))
    if not ((sequence is None and length == 0) or len(sequence) == length):
        raise ValueError(f'Part "{name}" has mismatched sequence length: check for bad characters and extra whitespace')

    # identity comes from source if set to a literal table, from display_id if not set
    identity = None
    if source_id and source_prefix:
        if source_prefix.strip() in source_table:
            display_id = string_to_display_id(source_id.strip())
            identity = f'{source_table[source_prefix.strip()]}/{display_id}'
        else:
            logging.warning(f'Part "{name}" ignoring non-literal source: {source_prefix}')
    elif source_id:
        logging.warning(f'Part "{name}" has source ID specified but not prefix: {source_id}')
    elif source_prefix:
        logging.warning(f'Part "{name}" has source prefix specified but not ID: {source_prefix}')
    if not identity:
        display_id = string_to_display_id(name)

    # build a component from the material
    logging.debug(f'Creating basic part "{name}"')
    component = sbol3.Component(identity or display_id, sbol3.SBO_DNA, name=name,
                                description=f'{design_notes}\n{description}'.strip())
    doc.add(component)
    if role:
        component.roles.append(role)
    if circular:
        component.types.append(sbol3.SO_CIRCULAR)
    if sequence:
        sbol_seq = sbol3.Sequence(f'{display_id}_sequence', encoding=sbol3.IUPAC_DNA_ENCODING, elements=sequence)
        doc.add(sbol_seq)
        component.sequences.append(sbol_seq.identity)

    # add the component to the appropriate collections
    basic_parts.members.append(component.identity)
    if final_product:
        linear_products.members.append(component.identity)
        final_products.members.append(component.identity)


##########################################
# Functions for parsing sub-components
# form of a sub-component:
# X: identifies a component or set thereof
# RC(X): X is reversed
reverse_complement_pattern = re.compile('RC\(.+\)')
# Returns sanitized text without optional reverse complement marker
def strip_RC(name):
    sanitized = name.strip()
    match = reverse_complement_pattern.match(sanitized)
    return (sanitized[3:-1] if (match and len(match.group())==len(sanitized)) else sanitized)
# returns true if part is reverse complement
def is_RC(name):
    sanitized = name.strip()
    return len(strip_RC(sanitized))<len(sanitized)
# returns a list of part names
def part_names(specification):
    return [name.strip() for name in strip_RC(str(specification)).split(',')]
# list all the parts in the row that aren't fully resolved
def unresolved_subparts(doc: sbol3.Document, row, config):
    return [name for spec in part_specifications(row, config) for name in part_names(spec) if not partname_to_part(doc,name)]
# get the part specifications until they stop
def part_specifications(row, config):
    return (cell.value for cell in row[config['composite_first_part_col']:] if cell.value)
def partname_to_part(doc: sbol3.Document, name_or_display_id: str):
    """Look up a part by its displayID or its name, searching first by displayID, then by name

    :param doc: SBOL document to search
    :param name_or_display_id: string to look up
    :return: object if found, None if not
    """
    return doc.find(name_or_display_id) or toplevel_named(doc,name_or_display_id)

###############################################################
# Functions for making composites, combinatorials, and libraries

def make_composite_component(display_id,part_lists,reverse_complements):
    # Make the composite as an engineered region
    composite_part = sbol3.Component(display_id, sbol3.SBO_DNA)
    composite_part.roles.append(sbol3.SO_ENGINEERED_REGION)
    # for each part, make a SubComponent and link them together in sequence
    last_sub = None
    for part_list,rc in zip(part_lists,reverse_complements):
        if not len(part_list)==1:
            raise ValueError(f'Part list should have precisely one element, but is {part_list}')
        sub = sbol3.SubComponent(part_list[0])
        sub.orientation = (sbol3.SBOL_REVERSE_COMPLEMENT if rc else sbol3.SBOL_INLINE)
        composite_part.features.append(sub)
        if last_sub: composite_part.constraints.append(sbol3.Constraint(sbol3.SBOL_MEETS,last_sub,sub))
        last_sub = sub
    # return the completed part
    return composite_part

constraint_pattern = re.compile('Part (\d+) (.+) Part (\d+)')
constraint_dict = {'same as': sbol3.SBOL_VERIFY_IDENTICAL,
                   'different from': sbol3.SBOL_DIFFERENT_FROM,
                   'same orientation as': sbol3.SBOL_SAME_ORIENTATION_AS,
                   'different orientation from': sbol3.SBOL_SAME_ORIENTATION_AS}
def make_constraint(constraint, part_list):
    m = constraint_pattern.match(constraint)
    if not m:
        raise ValueError(f'Constraint "{constraint}" does not match pattern "Part X relation Part Y"')
    try:
        restriction = constraint_dict[m.group(2)]
    except KeyError:
        raise ValueError(f'Do not recognize constraint relation "{restriction}"')
    x = int(m.group(1))
    y = int(m.group(3))
    if x is y:
        raise ValueError(f'A part cannot constrain itself: {constraint}')
    for n in [x,y]:
       if not (0 < n <= len(part_list)):
           raise ValueError(f'Part number "{str(n)}" is not between 1 and {len(part_list)}')
    return sbol3.Constraint(restriction, part_list[x-1], part_list[y-1])


def make_combinatorial_derivation(document, display_id,part_lists,reverse_complements,constraints):
    # Make the combinatorial derivation and its template
    template = sbol3.Component(display_id + "_template", sbol3.SBO_DNA)
    document.add(template)
    cd = sbol3.CombinatorialDerivation(display_id, template)
    cd.strategy = sbol3.SBOL_ENUMERATE
    # for each part, make a SubComponent or LocalSubComponent in the template and link them together in sequence
    template_part_list = []
    for part_list,rc in zip(part_lists,reverse_complements):
        # it's a variable if there are multiple values or if there's a single value that's a combinatorial derivation
        if len(part_list)>1 or not isinstance(part_list[0],sbol3.Component):
            sub = sbol3.LocalSubComponent({sbol3.SBO_DNA}) # make a template variable
            sub.name = "Part "+str(len(template_part_list)+1)
            template.features.append(sub)
            var = sbol3.VariableFeature(cardinality=sbol3.SBOL_ONE, variable=sub)
            cd.variable_features.append(var)
            # add all of the parts as variables
            for part in part_list:
                if isinstance(part,sbol3.Component): var.variants.append(part)
                elif isinstance(part,sbol3.CombinatorialDerivation): var.variant_derivations.append(part)
                else: raise ValueError("Don't know how to make library element for "+part.name+", a "+str(part))
        else: # otherwise it's a fixed element of the template
            sub = sbol3.SubComponent(part_list[0])
            template.features.append(sub)
        # in either case, orient and order the template elements
        sub.orientation = (sbol3.SBOL_REVERSE_COMPLEMENT if rc else sbol3.SBOL_INLINE)
        if template_part_list: template.constraints.append(sbol3.Constraint(sbol3.SBOL_MEETS,template_part_list[-1],sub))
        template_part_list.append(sub)
    # next, add all of the constraints to the template
    #template.constraints = (make_constraint(c.strip(),template_part_list) for c in (constraints.split(',') if constraints else [])) # impacted by pySBOL3 appending
    c_list = (make_constraint(c.strip(),template_part_list) for c in (constraints.split(',') if constraints else []))
    for c in c_list: template.constraints.append(c)
    # return the completed part
    return cd


def make_composite_part(document, row, composite_parts, linear_products, final_products, config):
    """
    Create a composite part from a row in the composites sheet
    :param document: Document to add parts to
    :param row: Excel row to be processed
    :param composite_parts: collection of parts to add to
    :param linear_products: collection of linear parts to add to
    :param final_products: collection of final parts to add to
    :param config: dictionary of sheet parsing configuration variables
    """
    # Parse material from sheet row
    name = row[config['composite_name_col']].value
    display_id = string_to_display_id(name)
    design_notes = (row[config['composite_notes_col']].value if row[config['composite_notes_col']].value else "")
    description = \
        (row[config['composite_description_col']].value if row[config['composite_description_col']].value else "")
    final_product = row[config['composite_final_col']].value  # boolean
    transformed_strain = row[config['composite_strain_col']].value if config['composite_strain_col'] else None
    backbone_or_locus_raw = row[config['composite_context_col']].value if config['composite_context_col'] else None
    backbone_or_locus = part_names(backbone_or_locus_raw) if backbone_or_locus_raw else []
    constraints = row[config['composite_constraints_col']].value if config['composite_constraints_col'] else None
    reverse_complements = [is_RC(spec) for spec in part_specifications(row,config)]
    part_lists = \
        [[partname_to_part(document, name) for name in part_names(spec)] for spec in part_specifications(row, config)]
    combinatorial = any(x for x in part_lists if len(x) > 1 or isinstance(x[0], sbol3.CombinatorialDerivation))

    # Build the composite
    logging.debug(f'Creating {"library" if combinatorial else "composite part"} "{name}"')
    linear_dna_display_id = (f'{display_id}_ins' if backbone_or_locus else display_id)
    if combinatorial:
        composite_part = make_combinatorial_derivation(document, linear_dna_display_id, part_lists, reverse_complements,
                                                       constraints)
    else:
        composite_part = make_composite_component(linear_dna_display_id, part_lists, reverse_complements)
    composite_part.name = (f'{name} insert' if backbone_or_locus else name)
    composite_part.description = f'{design_notes}\n{description}'.strip()

    # add the component to the appropriate collections
    document.add(composite_part)
    composite_parts.members.append(composite_part.identity)
    if final_product:
        linear_products.members.append(composite_part.identity)

    ###############
    # Consider strain and locus information
    if transformed_strain:
        warnings.warn("Not yet handling strain information: "+transformed_strain)
    if backbone_or_locus:
        # TODO: handle integration locuses as well as plasmid backbones
        backbones = [partname_to_part(document,name) for name in backbone_or_locus]
        if any(b is None for b in backbones):
            raise ValueError(f'Could not find specified backbone(s) "{backbone_or_locus}"')
        if any(tyto.SO.get_uri_by_term('plasmid') not in b.roles for b in backbones):
            raise ValueError(f'Specified backbones "{backbone_or_locus}" are not all plasmids')
        if combinatorial:
            logging.debug(f"Embedding library '{composite_part.name}' in plasmid backbone(s) '{backbone_or_locus}'")
            plasmid = sbol3.Component(f'{display_id}_template', sbol3.SBO_DNA)
            document.add(plasmid)
            part_sub = sbol3.LocalSubComponent([sbol3.SBO_DNA], name="Inserted Construct")
            plasmid.features.append(part_sub)
            plasmid_cd = sbol3.CombinatorialDerivation(display_id, plasmid, name=name)
            document.add(plasmid_cd)
            part_var = sbol3.VariableFeature(cardinality=sbol3.SBOL_ONE, variable=part_sub)
            plasmid_cd.variable_features.append(part_var)
            part_var.variant_derivations.append(composite_part)
            if final_product:
                final_products.members.append(plasmid_cd)
        else:
            if len(backbones) == 1:
                logging.debug(f'Embedding part "{composite_part.name}" in plasmid backbone "{backbone_or_locus}"')
                plasmid = sbol3.Component(display_id, sbol3.SBO_DNA, name=name)
                document.add(plasmid)
                part_sub = sbol3.SubComponent(composite_part)
                plasmid.features.append(part_sub)
                if final_product:
                    final_products.members += {plasmid}
            else:
                logging.debug(f'Embedding part "{composite_part.name}" in plasmid library "{backbone_or_locus}"')
                plasmid = sbol3.Component(f'{display_id}_template', sbol3.SBO_DNA)
                document.add(plasmid)
                part_sub = sbol3.SubComponent(composite_part)
                plasmid.features.append(part_sub)
                plasmid_cd = sbol3.CombinatorialDerivation(display_id, plasmid, name=name)
                document.add(plasmid_cd)
                if final_product:
                    final_products.members.append(plasmid_cd)

        if len(backbones) == 1:
            backbone_sub = sbol3.SubComponent(backbones[0])
            plasmid.features.append(backbone_sub)
        else:
            backbone_sub = sbol3.LocalSubComponent([sbol3.SBO_DNA])
            backbone_sub.name = "Vector"
            plasmid.features.append(backbone_sub)
            backbone_var = sbol3.VariableFeature(cardinality=sbol3.SBOL_ONE, variable=backbone_sub)
            plasmid_cd.variable_features.append(backbone_var)
            backbone_var.variants += backbones

        plasmid.constraints.append(sbol3.Constraint(sbol3.SBOL_MEETS, part_sub, backbone_sub))
        plasmid.constraints.append(sbol3.Constraint(sbol3.SBOL_MEETS, backbone_sub, part_sub))


def excel_to_sbol(wb: openpyxl.Workbook, config: dict = None) -> sbol3.Document:
    """
    Take an open Excel file, return an SBOL document
    :param wb: openpyxl pointer to an Excel file
    :param config: dictionary of sheet parsing configuration variables
    :return: Document containing all SBOL extracted from Excel sheet
    """
    config = expand_configuration(config)
    doc = sbol3.Document()

    logging.info('Reading metadata for collections')
    basic_parts, composite_parts, linear_products, final_products, source_table = read_metadata(wb, doc, config)

    logging.info('Reading basic parts')
    for row in wb[config['basic_sheet']].iter_rows(min_row=config['basic_first_row']):
        row_to_basic_part(doc, row, basic_parts, linear_products, final_products, config, source_table)
    logging.info(f'Created {len(basic_parts.members)} basic parts')

    logging.info('Reading composite parts and libraries')
    # first collect all rows with names
    pending_parts = [row for row in wb[config['composite_sheet']].iter_rows(min_row=config['composite_first_row'])
                     if row[config['composite_name_col']].value]
    while pending_parts:
        ready = [row for row in pending_parts if not unresolved_subparts(doc, row, config)]
        if not ready:
            raise ValueError("Could not resolve subparts" + ''.join(
                (f"\n in '{row[config['composite_name_col']].value}':" +
                 ''.join(f" '{x}'" for x in unresolved_subparts(doc, row, config)))
                for row in pending_parts))
        for row in ready:
            make_composite_part(doc, row, composite_parts, linear_products, final_products, config)
        pending_parts = [p for p in pending_parts if p not in ready]  # subtract parts from stable list
    logging.info(f'Created {len(composite_parts.members)} composite parts or libraries')

    logging.info(f'Count {len(basic_parts.members)} basic parts, {len(composite_parts.members)} composites/libraries')
    report = doc.validate()
    logging.info(f'Validation of document found {len(report.errors)} errors and {len(report.warnings)} warnings')
    return doc


type_to_standard_extension = {  # TODO: remove after resolution of pySBOL3/issues/244
    sbol3.SORTED_NTRIPLES: '.nt',
    sbol3.NTRIPLES: '.nt',
    sbol3.JSONLD: '.json',
    sbol3.RDF_XML: '.xml',
    sbol3.TURTLE: '.ttl'
}


def main():
    """
    Main wrapper: read from input file, invoke excel_to_sbol, then write to output file
    """
    parser = argparse.ArgumentParser()
    parser.add_argument('excel_file', help="Excel file used as input")
    parser.add_argument('-n', '--namespace', dest='namespace',
                        help="Namespace for Components in output file")
    parser.add_argument('-l', '--local', dest='local', default=None,
                        help="Local path for Components in output file")
    parser.add_argument('-o', '--output', dest='output_file', default='out',
                        help="Name of SBOL file to be written")
    parser.add_argument('-t', '--file-type', dest='file_type', default=sbol3.SORTED_NTRIPLES,
                        help="Name of SBOL file to output to (excluding type)")
    parser.add_argument('--verbose', '-v', dest='verbose', action='count', default=0,
                        help="Print running explanation of conversion process")
    args_dict = vars(parser.parse_args())

    # Extract arguments:
    verbosity = args_dict['verbose']
    log_level = logging.WARN if verbosity == 0 else logging.INFO if verbosity == 1 else logging.DEBUG
    logging.getLogger().setLevel(level=log_level)
    output_file = args_dict['output_file']
    file_type = args_dict['file_type']
    excel_file = args_dict['excel_file']
    outfile_name = output_file+type_to_standard_extension[file_type]
    sbol3.set_namespace(args_dict['namespace'])
    # TODO: unkludge after resolution of https://github.com/SynBioDex/pySBOL3/issues/288
    if args_dict['local']:
        sbol3.set_namespace(f"{args_dict['namespace']}/{args_dict['local']}")

    # Read file, convert, and write resulting document
    logging.info('Accessing Excel file '+excel_file)
    sbol_document = excel_to_sbol(openpyxl.load_workbook(excel_file, data_only=True))
    sbol_document.write(outfile_name, file_type)
    logging.info('SBOL file written to '+outfile_name)
