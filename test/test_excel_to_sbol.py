import unittest

import sbol_utilities.excel_to_sbol
import openpyxl
import tempfile
import sbol3
import os
import logging
import sys
from unittest.mock import patch

from test_helpers import assert_files_identical

TESTFILE_DIR = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'test_files')

logging.getLogger().setLevel(level=logging.DEBUG)


class TestExcel2SBOL(unittest.TestCase):
    def test_conversion(self):
        """Basic smoke test of Excel to SBOL3 conversion"""
        wb = openpyxl.load_workbook(os.path.join(TESTFILE_DIR, 'simple_library.xlsx'), data_only=True)
        sbol3.set_namespace('http://sbolstandard.org/testfiles/')
        doc = sbol_utilities.excel_to_sbol.excel_to_sbol(wb)

        assert not doc.validate().errors and not doc.validate().warnings
        assert len(doc.find('BasicParts').members) == 26
        assert len(doc.find('CompositeParts').members) == 6
        assert len(doc.find('LinearDNAProducts').members) == 2
        assert len(doc.find('FinalProducts').members) == 2

        temp_name = tempfile.mkstemp(suffix='.nt')[1]
        doc.write(temp_name, sbol3.SORTED_NTRIPLES)
        assert_files_identical(temp_name, os.path.join(TESTFILE_DIR, 'simple_library.nt'))

    def test_custom_conversion(self):
        """Test if conversion works correctly when the config us used to change expected sheet structure"""
        wb = openpyxl.load_workbook(os.path.join(TESTFILE_DIR, 'nonstandard_simple_library.xlsx'), data_only=True)
        sbol3.set_namespace('http://sbolstandard.org/testfiles/')
        config = {
            'basic_parts_name': 'C2',
            'basic_parts_description': 'A12',
            'basic_first_row': 21,
            'basic_role_col': 2,
            'basic_notes_col': 3,
            'basic_description_col': 5,
            'basic_source_prefix_col': 6,
            'basic_source_id_col': 7,
            'basic_final_col': 10,
            'basic_circular_col': 11,
            'basic_length_col': 12,
            'basic_sequence_col': 13,
            'composite_first_part_col': 8
        }
        doc = sbol_utilities.excel_to_sbol.excel_to_sbol(wb, config)

        assert not doc.validate().errors and not doc.validate().warnings
        assert len(doc.find('BasicParts').members) == 26
        assert len(doc.find('CompositeParts').members) == 6
        assert len(doc.find('LinearDNAProducts').members) == 2
        assert len(doc.find('FinalProducts').members) == 2

        temp_name = tempfile.mkstemp(suffix='.nt')[1]
        doc.write(temp_name, sbol3.SORTED_NTRIPLES)
        assert_files_identical(temp_name, os.path.join(TESTFILE_DIR, 'simple_library.nt'))

    def test_multi_backbone(self):
        """Check if generation works correclty when there is more than one backbone option"""
        wb = openpyxl.load_workbook(os.path.join(TESTFILE_DIR, 'two_backbones.xlsx'), data_only=True)
        sbol3.set_namespace('http://sbolstandard.org/testfiles/')
        doc = sbol_utilities.excel_to_sbol.excel_to_sbol(wb)
        assert not doc.validate().errors and not doc.validate().warnings
        assert len(doc.find('BasicParts').members) == 9
        assert len(doc.find('CompositeParts').members) == 2
        assert len(doc.find('LinearDNAProducts').members) == 2
        assert len(doc.find('FinalProducts').members) == 2

        temp_name = tempfile.mkstemp(suffix='.nt')[1]
        doc.write(temp_name, sbol3.SORTED_NTRIPLES)
        assert_files_identical(temp_name, os.path.join(TESTFILE_DIR, 'two_backbones.nt'))

    def test_constraints(self):
        """Check if constraints are generated correctly"""
        wb = openpyxl.load_workbook(os.path.join(TESTFILE_DIR, 'constraints_library.xlsx'), data_only=True)
        sbol3.set_namespace('http://sbolstandard.org/testfiles/')
        doc = sbol_utilities.excel_to_sbol.excel_to_sbol(wb)

        assert not doc.validate().errors and not doc.validate().warnings
        assert len(doc.find('BasicParts').members) == 43
        assert len(doc.find('CompositeParts').members) == 8
        assert len(doc.find('LinearDNAProducts').members) == 2
        assert len(doc.find('FinalProducts').members) == 2

        temp_name = tempfile.mkstemp(suffix='.nt')[1]
        doc.write(temp_name, sbol3.SORTED_NTRIPLES)
        assert_files_identical(temp_name, os.path.join(TESTFILE_DIR, 'constraints_library.nt'))

    def test_commandline(self):
        """Make sure function works correctly when run from the command line"""
        temp_name = tempfile.mkstemp(suffix='.nt')[1]
        test_args = ['excel_file', '-vv', os.path.join(TESTFILE_DIR, 'simple_library.xlsx'), '-o', temp_name, '-n',
                     'http://sbolstandard.org/testfiles/']
        with patch.object(sys, 'argv', test_args):
            sbol_utilities.excel_to_sbol.main()
        assert_files_identical(temp_name, os.path.join(TESTFILE_DIR, 'simple_library.nt'))

if __name__ == '__main__':
    unittest.main()
