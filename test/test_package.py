import unittest
import os
from contextlib import contextmanager
from pathlib import Path
import tempfile
import shutil

import openpyxl
import sbol3

from sbol_utilities import package, excel_to_sbol
from sbol_utilities.helper_functions import sbol3_namespace, member_named, reference_named
from sbol_utilities.package import PackageError, sep_054
from sbol_utilities.sbol_diff import doc_diff, file_diff

TEST_FILES = Path(__file__).parent / 'test_files'


@contextmanager
def temporary_package_manager() -> package.PackageManager:
    """For ease of testing, create a context manager for clean package loading environments

    :return: temporary package manager
    """
    saved = package.ACTIVE_PACKAGE_MANAGER
    with tempfile.TemporaryDirectory() as temp_dir_name:
        tempdir = Path(temp_dir_name)
        package.ACTIVE_PACKAGE_MANAGER = package.PackageManager(catalog_directory=tempdir)
        yield package.ACTIVE_PACKAGE_MANAGER
    package.ACTIVE_PACKAGE_MANAGER = saved


class TestPackage(unittest.TestCase):
    def test_define_package_single_file(self):
        """Test defining a package from an SBOL document"""
        # Read in the test file
        doc_01 = sbol3.Document()
        doc_01.read(str(TEST_FILES / 'package_in_01.nt'))

        # Create the package and add it to a document
        out_doc = sbol3.Document()
        out_01 = package.doc_to_package(doc_01)
        out_doc.add(out_01)
        self.assertTrue(not out_doc.validate().errors and not out_doc.validate().warnings)

        # Compare it to the saved results file, make sure they are the same
        expected = sbol3.Document()
        expected.read(str(TEST_FILES / 'package_out_01.nt'))
        self.assertFalse(doc_diff(out_doc, expected))

    def test_define_package_multi_file(self):
        """Test defining a package from multiple SBOL documents"""
        # Read in the test files
        test_dir = os.path.dirname(os.path.realpath(__file__))
        doc_01 = sbol3.Document()
        doc_01.read(os.path.join(test_dir, 'test_files', 'package_in_01.nt'))

        doc_02 = sbol3.Document()
        doc_02.read(os.path.join(test_dir, 'test_files', 'package_in_02.nt'))

        doc_03 = sbol3.Document()
        doc_03.read(os.path.join(test_dir, 'test_files', 'package_in_03.nt'))

        # Run the function
        # Here, I only want the package object, not any of the subpackages
        out_02 = package.docs_to_package(doc_01, [doc_02, doc_03])

        # Write a temporary file
        doc = sbol3.Document()
        doc.add(out_02)
        self.assertTrue(not doc.validate().errors and not doc.validate().warnings)
        tmp_out = tempfile.mkstemp(suffix='.nt')[1]
        doc.write(tmp_out, sbol3.SORTED_NTRIPLES)

        # Compare it to the saved results file, make sure they are the same
        comparison_file = os.path.join(test_dir, 'test_files', 'package_out_02.nt')
        self.assertFalse(file_diff(tmp_out, comparison_file))

    def test_prefix_too_short(self):
        """ Test that having a sub-package with a different namespace than the 
        root package fails. All members of "package_in_01_error.nt" has the
        namespace "https://bad-example.org/MyPackage", whereas the subpackages
        are "https://example.org/MyPackage/promoters" and
        "https://example.org/MyPackage/repressors". Will raise a value error
        in check_prefix that "'The packages {root_package} and {sub_package}
        namespace URIs do not share the same URL scheme specifier or network
        location, and so do not represent a root and sub package.'
        """
        # Read in the test files
        test_dir = os.path.dirname(os.path.realpath(__file__))
        doc_01 = sbol3.Document()
        doc_01.read(os.path.join(test_dir, 'test_files', 'package_in_01_error.nt'))

        doc_02 = sbol3.Document()
        doc_02.read(os.path.join(test_dir, 'test_files', 'package_in_02.nt'))

        doc_03 = sbol3.Document()
        doc_03.read(os.path.join(test_dir, 'test_files', 'package_in_03.nt'))

        # Run the function
        with self.assertRaises(ValueError):
            package.docs_to_package(doc_01, [doc_02, doc_03])

    def test_subpackage_fails(self):
        """ Test that having one subpackage file with multiple namespaces fails.
        Raises value in define_package that not all members in the package have 
        the same namespace. "package_in_02_error.nt" has members with two 
        different namespaces, https://example.org/MyPackage/promoters and 
        https://example.org/MyPackage/inhibitors."""
        # Read in the test files
        test_dir = os.path.dirname(os.path.realpath(__file__))
        doc_01 = sbol3.Document()
        doc_01.read(os.path.join(test_dir, 'test_files', 'package_in_01.nt'))

        doc_02 = sbol3.Document()
        doc_02.read(os.path.join(test_dir, 'test_files', 'package_in_02_error.nt'))

        doc_03 = sbol3.Document()
        doc_03.read(os.path.join(test_dir, 'test_files', 'package_in_03.nt'))

        # Run the function
        with self.assertRaises(ValueError):
            package.docs_to_package(doc_01, [doc_02, doc_03])

    def test_make_package_from_MyPackage(self):
        """ Create a package based on the files in a directory (test/test_files/
        MyPackage). The function automatically saves the package files in the 
        .sip package directories of each subdirectory. """
        # Set the package directory
        test_dir = os.path.dirname(os.path.realpath(__file__))
        dir_name = os.path.join(test_dir, 'test_files', 'MyPackage')

        # Pass to the function
        package.directory_to_package(dir_name)

        # Compare all the package files to the saved results file, make sure
        # they are the same, then delete the package directory
        for root, _, _ in os.walk(dir_name):
            # Collect the output from the actual function
            # Want to get the path separate from the file for easy deleting
            out_path = os.path.join(root, '.sip')
            out_file = os.path.join(out_path, 'package.nt')

            # I have saved all the results to compare against in one
            # directory with the names corresponding to the subdirectory name
            # from the original directory
            file_name = root.split('/')[-1] + '.nt'
            comparison_file = os.path.join(test_dir,
                                           'test_files',
                                           'MyPackage-results',
                                           file_name)

            self.assertFalse(file_diff(out_file, comparison_file))

            # Delete the package directory
            # TODO: change to a temp directory so we don't have to do this
            shutil.rmtree(out_path, ignore_errors=True)

    def test_make_package_from_MyPackage_w_multiple_files(self):
        """ Create a package based on the files in a directory (test/test_files/
        MyPackage_w_multiple_files). The function automatically saves the 
        package file in the .sip package directories of each subdirectory. """
        # Set the package directory
        test_dir = os.path.dirname(os.path.realpath(__file__))
        dir_name = os.path.join(test_dir,
                                'test_files',
                                'MyPackage_w_multiple_files')

        # Pass to the function
        package.directory_to_package(dir_name)

        # Compare all the package files to the saved results file, make sure
        # they are the same, then delete the package directory
        for root, _, _ in os.walk(dir_name):
            # Collect the output from the actual function
            # Want to get the path separate from the file for easy deleting
            out_path = os.path.join(root, '.sip')
            out_file = os.path.join(out_path, 'package.nt')

            # I have saved all the results to compare against in one
            # directory with the names corresponding to the subdirectory name
            # from the original directory
            file_name = root.split('/')[-1] + '.nt'
            comparison_file = os.path.join(test_dir,
                                           'test_files',
                                           'MyPackage_w_multiple_files-results',
                                           file_name)

            self.assertFalse(file_diff(out_file, comparison_file))

            # Delete the package directory
            # TODO: change to a temp directory so we don't have to do this
            shutil.rmtree(out_path, ignore_errors=True)

    def test_make_package_from_MyPackage_w_sub_sub_packages(self):
        """ Create a package based on the files in a directory (test/test_files/
        MyPackage_w_sub_sub_packages). The function automatically saves the 
        package file in the .sip package directories of each subdirectory. """
        # Set the package directory
        test_dir = os.path.dirname(os.path.realpath(__file__))
        dir_name = os.path.join(test_dir,
                                'test_files',
                                'MyPackage_w_sub_sub_packages')

        # Pass to the function
        package.directory_to_package(dir_name)

        # Compare all the package files to the saved results file, make sure
        # they are the same, then delete the package directory
        for root, _, _ in os.walk(dir_name):
            # Collect the output from the actual function
            # Want to get the path separate from the file for easy deleting
            out_path = os.path.join(root, '.sip')
            out_file = os.path.join(out_path, 'package.nt')

            # I have saved all the results to compare against in one
            # directory with the names corresponding to the subdirectory name
            # from the original directory
            file_name = root.split('/')[-1] + '.nt'
            comparison_file = os.path.join(test_dir,
                                           'test_files',
                                           'MyPackage_w_sub_sub_packages-results',
                                           file_name)

            self.assertFalse(file_diff(out_file, comparison_file))

            # Delete the package directory
            # TODO: change to a temp directory so we don't have to do this
            shutil.rmtree(out_path, ignore_errors=True)

    def test_install_package(self):
        """Test installation, loading, and validation of a package"""
        # create a temporary package manager for testing
        with temporary_package_manager() as pm:
            # make and install a miniature package
            doc = sbol3.Document()
            doc.read(str(TEST_FILES / 'BBa_J23101_package_namespace.nt'))
            p = package.doc_to_package(doc)
            package.install_package(p.namespace, doc)
            # check that package manager has the right contents
            self.assertEqual(len(pm.package_catalog), 1)
            self.assertTrue('https://synbiohub.org/public/igem' in pm.package_catalog)
            # check that the files that are created look like what is expected
            catalog_dir = pm._catalog_directory
            _HASH_NAME = 'c15f6ed6923d64b7516b632b6c32ff93.nt'
            self.assertEqual(2, len(list(catalog_dir.iterdir())))
            self.assertFalse(file_diff(TEST_FILES / 'BBa_J23101_package.nt', catalog_dir / _HASH_NAME))
            catalog_doc = sbol3.Document()
            catalog_doc.read(str(catalog_dir / 'package-catalog.nt'))
            self.assertEqual(2, len(catalog_doc.objects))
            p = catalog_doc.find('https://synbiohub.org/public/igem/package')
            self.assertEqual((catalog_dir / _HASH_NAME).as_uri(), p.attachments[0].lookup().source)

    def test_fat_package(self):
        """Test installation, loading, and validation of a package that contains sub-packages and dependencies"""
        with temporary_package_manager():  # create a temporary package manager for testing
            # make and install a package with a sub-package & dependency
            doc = sbol3.Document()
            doc.read(str(TEST_FILES / 'MyPackage' / 'package_in_01.nt'))
            p = package.doc_to_package(doc)
            # add the sub-package
            doc2 = sbol3.Document()
            doc2.read(str(TEST_FILES / 'MyPackage' / 'promoters' / 'package_in_02.nt'))
            p2 = package.doc_to_package(doc2)
            p.subpackages.append(p2)
            sbol3.copy(doc2.objects, into_document=doc)  # fatten the package with included sub-package
            # add a dependency with a different namespace
            doc3 = sbol3.Document()
            doc3.read(str(TEST_FILES / 'BBa_J23101_package_namespace.nt'))
            p3 = package.doc_to_package(doc3)
            p.dependencies.append(sep_054.Dependency(package=p3))
            sbol3.copy(doc3.objects, into_document=doc)  # fatten the package with included dependency
            package.install_package(p.namespace, doc)

            # check that we can load the package and get materials from both package and subpackage
            package.load_package(p.namespace)
            self.assertIsNotNone(package.lookup('https://example.org/MyPackage/E0040'))
            self.assertIsNotNone(package.lookup('https://example.org/MyPackage/promoters/J364007'))
            self.assertIsNotNone(package.lookup('https://synbiohub.org/public/igem/BBa_J23101'))

            # Make sure we catch embedded package conflicts
            doc4 = sbol3.Document()
            doc4.read(str(TEST_FILES / 'MyPackage' / 'repressors' / 'package_in_03.nt'))
            p4 = package.doc_to_package(doc4)
            # add the same dependency
            p4.dependencies.append(sep_054.Dependency(package=p3))
            sbol3.copy(doc3.objects, into_document=doc4)  # fatten the package with duplicate dependency
            package.install_package(p4.namespace, doc4)

            with self.assertRaises(PackageError) as cm:
                package.load_package(p4.namespace)
            msg = 'Embedded package would override already-loaded package: https://synbiohub.org/public/igem'
            self.assertTrue(str(cm.exception).startswith(msg))

    def test_dissociated_packages(self):
        """Test installation, loading, and validation of dissociated packages"""
        with temporary_package_manager() as pm:  # create a temporary package manager for testing
            # install two packages containing dissociated fragments
            doc = sbol3.Document()
            doc.read(str(TEST_FILES / 'BBa_J23101_package_dissociated.nt'))
            package.install_package('https://synbiohub.org/public/igem', doc)
            doc.read(str(TEST_FILES / 'dissociated_dependencies.nt'))
            package.install_package('https://test.org/dissociated', doc)

            # check that package manager has the right contents (i.e., roots only)
            self.assertEqual(len(pm.package_catalog), 2)
            self.assertTrue('https://synbiohub.org/public/igem' in pm.package_catalog)
            self.assertTrue('https://test.org/dissociated' in pm.package_catalog)
            # load the packages to make sure they can enter as expected
            p1 = package.load_package('https://test.org/dissociated')
            self.assertIsNone(package.lookup('https://synbiohub.org/public/igem/BBa_J23101'))  # in the second fragment
            c1 = package.lookup('https://www.ncbi.nlm.nih.gov/nuccore/ATP07149_1')
            self.assertEqual(type(c1), sbol3.Component)
            self.assertEqual(c1.document, p1.document)
            c2 = package.lookup('https://synbiohub.org/public/igem/E0040')
            self.assertEqual(type(c2), sbol3.Component)
            self.assertEqual(c2.document, p1.document)
            p2 = package.load_package('https://synbiohub.org/public/igem')
            self.assertNotEqual(p1, p2)  # should have loaded a new fragment
            c3 = package.lookup('https://synbiohub.org/public/igem/BBa_J23101')
            self.assertEqual(type(c3), sbol3.Component)
            self.assertEqual(c3.document, p2.document)
            self.assertEqual(c1, package.lookup('https://www.ncbi.nlm.nih.gov/nuccore/ATP07149_1'))
            self.assertEqual(c2, package.lookup('https://synbiohub.org/public/igem/E0040'))
            # check idempotence of loading
            self.assertEqual(p1, package.load_package('https://test.org/dissociated'))
            self.assertEqual(p2, package.load_package('https://synbiohub.org/public/igem'))
            p3 = package.load_package('https://synbiohub.org/public/igem', doc=p1.document)  # embedded reload
            self.assertEqual(p3.document, p1.document)  # should get the existing embedded package

    def test_package_validation(self):
        """Test that package system can load packages and verify integrity of the load"""
        with temporary_package_manager():
            # Tried to load a non-existent file
            with self.assertRaises(PackageError) as cm:
                package.load_package('https://synbiohub.org/public/', TEST_FILES / 'this_file_does_not_exist.nt')
            wrap_msg = 'Could not read package https://synbiohub.org/public/'
            self.assertTrue(str(cm.exception).startswith(wrap_msg))
            self.assertTrue(isinstance(cm.exception.__context__, FileNotFoundError))

            # Does not contain specified package
            with self.assertRaises(PackageError) as cm:
                package.load_package('https://synbiohub.org/public/', TEST_FILES / 'BBa_J23101_package.nt')
            wrap_msg = 'Validation error while loading package https://synbiohub.org/public/'
            self.assertTrue(str(cm.exception).startswith(wrap_msg))
            self.assertTrue(str(cm.exception.__context__).startswith('Cannot find Package in SBOL document'))

            # object isn't a package
            with self.assertRaises(PackageError) as cm:
                package.load_package('https://synbiohub.org/public/BBa_J23101', TEST_FILES/'BBa_J23101_bad_package.nt')
            self.assertTrue(str(cm.exception).startswith(wrap_msg))
            msg = 'Object should be a Package, but is not: <Component https://synbiohub.org/public/BBa_J23101/package>'
            self.assertTrue(str(cm.exception.__context__).startswith(msg))

            # Package has the wrong namespace
            with self.assertRaises(PackageError) as cm:
                package.load_package('https://synbiohub.org/public/igem', TEST_FILES / 'BBa_J23101_bad_package.nt')
            self.assertTrue(str(cm.exception).startswith(wrap_msg))
            msg = 'Package https://synbiohub.org/public/igem/package should have namespace ' \
                  'https://synbiohub.org/public/igem but was https://synbiohub.org'
            self.assertTrue(str(cm.exception.__context__).startswith(msg))

            # Package is missing members
            with self.assertRaises(PackageError) as cm:
                package.load_package('https://synbiohub.org/public/igem2', TEST_FILES / 'BBa_J23101_bad_package.nt')
            self.assertTrue(str(cm.exception).startswith(wrap_msg))
            msg = 'Package https://synbiohub.org/public/igem2 missing listed members: ' \
                  '[\'https://synbiohub.org/public/igem/BBa_J23101_not_here\']'
            self.assertTrue(str(cm.exception.__context__).startswith(msg))

            # Package has unaccounted for objects
            with self.assertRaises(PackageError) as cm:
                package.load_package('https://synbiohub.org/public/igem3', TEST_FILES / 'BBa_J23101_bad_package2.nt')
            self.assertTrue(str(cm.exception).startswith(wrap_msg))
            msg = 'Package https://synbiohub.org/public/igem3 contains unexpected TopLevel objects: ' \
                  '[\'https://synbiohub.org/public/BBa_J23101\']'
            self.assertTrue(str(cm.exception.__context__).startswith(msg))

            # Package document has unexpected embedded packages
            with self.assertRaises(PackageError) as cm:
                package.load_package('https://synbiohub.org/public/igem2', TEST_FILES / 'BBa_J23101_bad_package3.nt')
            self.assertTrue(str(cm.exception).startswith(wrap_msg))
            msg = 'Document embeds packages not referred to from root: [\'https://synbiohub.org/public/igem/package\'' \
                  ', \'https://synbiohub.org/public/igem3/package\']'
            self.assertTrue(str(cm.exception.__context__).startswith(msg))

    def test_cross_document_lookup(self):
        """Test that package system correctly overrides the document lookup function to enable cross-package lookups"""
        with temporary_package_manager():
            p = package.load_package('https://synbiohub.org/public/igem', TEST_FILES / 'BBa_J23101_package.nt')
            # make sure loading is idempotent (i.e., it returns the existing load rather than reloading)
            self.assertEqual(p, package.load_package('https://synbiohub.org/public/igem'))
            doc = sbol3.Document()
            good_id = 'https://synbiohub.org/public/igem/BBa_J23101_sequence'
            bad_id = 'https://synbiohub.org/public/igem/BBa_J23101_notexist'
            bad_package = 'https://badpackage/not_exist'
            with sbol3_namespace('http://foo.bar/baz'):
                doc.add(sbol3.Component('qux', sbol3.SBO_DNA, sequences=[good_id, bad_id, bad_package]))

            c = doc.find('qux')
            # check for successful cross-document lookup
            self.assertEqual(good_id, c.sequences[0].lookup().identity)
            self.assertEqual(good_id, package.lookup(c.sequences[0]).identity)
            # check for appropriate handling of failed lookups
            self.assertEqual(None, c.sequences[1].lookup())
            self.assertEqual(None, package.lookup(bad_id))
            with self.assertRaises(PackageError) as cm:
                c.sequences[2].lookup()
            self.assertTrue(str(cm.exception).startswith('No loaded package has a namespace that contains'))
        # make sure temporary package manager operated cleanly
        p2 = package.load_package('https://synbiohub.org/public/igem', TEST_FILES / 'BBa_J23101_package.nt')
        self.assertNotEqual(p, p2)

    def test_packaged_excel(self):
        """Basic smoke test of Excel to SBOL3 conversion with a package"""
        wb = openpyxl.load_workbook(TEST_FILES / 'packaged_library_no_dissociated.xlsx', data_only=True)
        doc = excel_to_sbol.excel_to_sbol(wb)
        # Make sure package is valid
        self.assertTrue(not doc.validate().errors and not doc.validate().warnings)
        # check that the package has all the expected members
        # 17 basic parts + 15 sequences + 1 composites + 1x4 CombDev/BB (4)c + 4 product collections = 41
        self.assertEqual(41, len(doc.find('package').members))

        # Write it out and make sure we can import as a package
        temp_name = tempfile.mkstemp(suffix='.nt')[1]
        doc.write(temp_name, sbol3.SORTED_NTRIPLES)
        with temporary_package_manager():
            package.load_package('https://test.org/mypackage', temp_name)

    # def test_dissociated_package_excel(self):
    #     """Basic smoke test of Excel to SBOL3 conversion with a package"""
    #     wb = openpyxl.load_workbook(TEST_FILES / 'packaged_library.xlsx', data_only=True)
    #     doc = excel_to_sbol.excel_to_sbol(wb)
    #     # Make sure package is valid
    #     self.assertTrue(not doc.validate().errors and not doc.validate().warnings)
    #     # check that the package has all the expected members
    #     # 17 basic parts + 15 sequences + 1 composites + 1x4 CombDev/BB (4)c + 4 product collections = 41
    #     self.assertEqual(41, len(doc.find('package').members))
    #
    #     # Write it out and make sure we can import as a package
    #     temp_name = tempfile.mkstemp(suffix='.nt')[1]
    #     doc.write(temp_name, sbol3.SORTED_NTRIPLES)
    #     with temporary_package_manager():
    #         package.load_package('https://test.org/mypackage', temp_name)

    def test_excel_imports(self):
        """Excel to SBOL3 conversion including a package dependency"""
        with temporary_package_manager():
            # make sure package with dependency will refuse to convert until dependency is installed
            wb = openpyxl.load_workbook(TEST_FILES / 'second_library.xlsx', data_only=True)
            with self.assertRaises(PackageError) as cm:
                doc = excel_to_sbol.excel_to_sbol(wb)

            # convert base package:
            wb = openpyxl.load_workbook(TEST_FILES / 'packaged_library_no_dissociated.xlsx', data_only=True)
            doc = excel_to_sbol.excel_to_sbol(wb)
            # Install package, allowing it to be implicitly loaded
            package.install_package('https://test.org/mypackage/', doc)

            # convert package with dependency
            wb = openpyxl.load_workbook(TEST_FILES / 'second_library.xlsx', data_only=True)
            doc = excel_to_sbol.excel_to_sbol(wb)
            # Make sure package is valid
            self.assertTrue(not doc.validate().errors and not doc.validate().warnings)
            # check that the package has all the expected members
            # 7 basic parts + 7 sequences + 1x2 CombDev + 4 product collections = 20
            self.assertEqual(20, len(doc.find('package').members))
            # Check that the package has the expected dependency
            self.assertEqual(['https://test.org/mypackage'], [str(d.package) for d in doc.find('package').dependencies])

            # Write it out and make sure we can import as a package
            temp_name = tempfile.mkstemp(suffix='.nt')[1]
            doc.write(temp_name, sbol3.SORTED_NTRIPLES)
            second_library = package.load_package('https://test.org/second_library', temp_name)

            # Check that cross-document link works correctly
            fp_collection = member_named(second_library, 'All FPs')
            self.assertTrue(isinstance(fp_collection, sbol3.CombinatorialDerivation))
            rfp_component = reference_named(fp_collection.variable_features[0].variants, 'mRFP1')
            self.assertIsNotNone(rfp_component)
            self.assertEqual(rfp_component.identity, 'https://test.org/mypackage/mRFP1')
            self.assertEqual(rfp_component.description, 'Red FP (off patent)\nmRFP1')


if __name__ == '__main__':
    unittest.main()
